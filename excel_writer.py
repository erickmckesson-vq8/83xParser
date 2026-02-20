"""Generate readable Excel files from parsed EDI data."""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# Styling constants
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="2B5797", end_color="2B5797", fill_type="solid")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
CURRENCY_FORMAT = '#,##0.00'
THIN_BORDER = Border(
    left=Side(style="thin", color="D0D0D0"),
    right=Side(style="thin", color="D0D0D0"),
    top=Side(style="thin", color="D0D0D0"),
    bottom=Side(style="thin", color="D0D0D0"),
)
ALT_ROW_FILL = PatternFill(start_color="EBF1F8", end_color="EBF1F8", fill_type="solid")


def style_header(ws, num_cols):
    """Apply header styling to the first row."""
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER


def style_data(ws, num_rows, num_cols, currency_cols=None):
    """Apply data styling: alternating rows, borders, currency formatting."""
    currency_cols = currency_cols or []
    for row in range(2, num_rows + 1):
        for col in range(1, num_cols + 1):
            cell = ws.cell(row=row, column=col)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="center")
            if (row % 2) == 0:
                cell.fill = ALT_ROW_FILL
            if col in currency_cols:
                cell.number_format = CURRENCY_FORMAT


def auto_width(ws, num_cols, max_width=50):
    """Auto-size column widths based on content."""
    for col in range(1, num_cols + 1):
        max_len = 0
        for row in ws.iter_rows(min_col=col, max_col=col, values_only=False):
            for cell in row:
                val = str(cell.value) if cell.value is not None else ""
                max_len = max(max_len, len(val))
        adjusted = min(max_len + 3, max_width)
        ws.column_dimensions[get_column_letter(col)].width = max(adjusted, 10)


def write_sheet(ws, headers, rows, currency_cols=None):
    """Write a complete sheet with headers, data, and styling."""
    # Write headers
    for col_idx, header in enumerate(headers, 1):
        ws.cell(row=1, column=col_idx, value=header)

    # Write data
    for row_idx, row_data in enumerate(rows, 2):
        for col_idx, value in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    num_cols = len(headers)
    num_rows = len(rows) + 1
    style_header(ws, num_cols)
    style_data(ws, num_rows, num_cols, currency_cols)
    auto_width(ws, num_cols)
    ws.freeze_panes = "A2"


# ---------------------------------------------------------------------------
# Combined Excel Writer — handles multiple 835 and 837 files in one workbook
# ---------------------------------------------------------------------------

def write_combined_excel(all_835, all_837, output_path):
    """Write combined 835 and 837 data into a single Excel workbook.

    Args:
        all_835: list of (filename, parsed_transactions) tuples for 835 files
        all_837: list of (filename, parsed_transactions) tuples for 837 files
        output_path: file path for the output Excel file
    """
    wb = Workbook()
    first_sheet = True  # Track so we can rename the default sheet

    # -------------------------------------------------------------------
    # 835 sheets
    # -------------------------------------------------------------------
    if all_835:
        # --- 835 Payment Summary ---
        if first_sheet:
            ws_payment = wb.active
            ws_payment.title = "835 Payment Summary"
            first_sheet = False
        else:
            ws_payment = wb.create_sheet("835 Payment Summary")

        payment_headers = [
            "Source File",
            "Trace Number", "Payment Amount", "Payment Method",
            "Payment Date", "Payer Name", "Payer ID",
            "Payee Name", "Payee ID",
        ]
        payment_rows = []
        for filename, transactions in all_835:
            for txn in transactions:
                p = txn["payment"]
                payment_rows.append([
                    filename,
                    p["trace_number"], p["payment_amount"], p["payment_method"],
                    p["payment_date"], p["payer_name"], p["payer_id"],
                    p["payee_name"], p["payee_id"],
                ])
        write_sheet(ws_payment, payment_headers, payment_rows, currency_cols=[3])

        # --- 835 Claims ---
        ws_claims = wb.create_sheet("835 Claims")
        claim_headers = [
            "Source File",
            "Trace Number", "Patient Control Number", "Claim Status",
            "Patient Name", "Insured Name",
            "Total Charges", "Payment Amount", "Patient Responsibility",
            "Payer Claim Number", "Filing Indicator",
            "Rendering Provider", "Rendering Provider NPI",
            "Statement From", "Statement To",
            "Claim Received Date",
            "Total Adjustments (CO)", "Total Adjustments (PR)",
            "Total Adjustments (OA)",
        ]
        claim_rows = []
        for filename, transactions in all_835:
            for txn in transactions:
                trace = txn["payment"]["trace_number"]
                for c in txn["claims"]:
                    patient_name = f"{c['patient_last_name']}, {c['patient_first_name']}".strip(", ")
                    insured_name = f"{c['insured_last_name']}, {c['insured_first_name']}".strip(", ")
                    rendering = f"{c['rendering_provider_last_name']}, {c['rendering_provider_first_name']}".strip(", ")

                    adj_co = sum(a["amount"] for a in c["adjustments"] if a["group_code"] == "CO")
                    adj_pr = sum(a["amount"] for a in c["adjustments"] if a["group_code"] == "PR")
                    adj_oa = sum(a["amount"] for a in c["adjustments"]
                                 if a["group_code"] not in ("CO", "PR"))
                    for svc in c["service_lines"]:
                        adj_co += sum(a["amount"] for a in svc["adjustments"] if a["group_code"] == "CO")
                        adj_pr += sum(a["amount"] for a in svc["adjustments"] if a["group_code"] == "PR")
                        adj_oa += sum(a["amount"] for a in svc["adjustments"]
                                      if a["group_code"] not in ("CO", "PR"))

                    claim_rows.append([
                        filename,
                        trace, c["patient_control_number"], c["claim_status"],
                        patient_name, insured_name,
                        c["total_charge"], c["payment_amount"], c["patient_responsibility"],
                        c["payer_claim_number"], c["filing_indicator"],
                        rendering, c["rendering_provider_npi"],
                        c["claim_statement_from"], c["claim_statement_to"],
                        c["claim_received_date"],
                        adj_co, adj_pr, adj_oa,
                    ])
        write_sheet(ws_claims, claim_headers, claim_rows,
                    currency_cols=[7, 8, 9, 17, 18, 19])

        # --- 835 Service Lines ---
        ws_svc = wb.create_sheet("835 Service Lines")
        svc_headers = [
            "Source File",
            "Patient Control Number", "Procedure Code", "Modifiers",
            "Revenue Code", "Charge Amount", "Payment Amount",
            "Units Paid", "Service Date",
            "Original Procedure Code", "Original Units",
            "Adjustment Groups", "Remark Codes",
        ]
        svc_rows = []
        for filename, transactions in all_835:
            for txn in transactions:
                for c in txn["claims"]:
                    for svc in c["service_lines"]:
                        adj_summary = "; ".join(
                            f"{a['group_code']}-{a['reason_code']}: ${a['amount']:.2f}"
                            for a in svc["adjustments"]
                        )
                        remarks = ", ".join(svc.get("remark_codes", []))
                        svc_rows.append([
                            filename,
                            svc["_claim_id"], svc["procedure_code"], svc["modifiers"],
                            svc["revenue_code"], svc["charge_amount"], svc["payment_amount"],
                            svc["units_paid"], svc["service_date"],
                            svc["original_procedure_code"], svc["original_units"],
                            adj_summary, remarks,
                        ])
        write_sheet(ws_svc, svc_headers, svc_rows, currency_cols=[6, 7])

        # --- 835 Adjustments ---
        ws_adj = wb.create_sheet("835 Adjustments")
        adj_headers = [
            "Source File",
            "Patient Control Number", "Level",
            "Group Code", "Group Description",
            "Reason Code", "Reason Description",
            "Adjustment Amount", "Quantity",
        ]
        adj_rows = []
        for filename, transactions in all_835:
            for txn in transactions:
                for c in txn["claims"]:
                    for adj in c["adjustments"]:
                        adj_rows.append([
                            filename,
                            c["patient_control_number"], "Claim",
                            adj["group_code"], adj["group_description"],
                            adj["reason_code"], adj["reason_description"],
                            adj["amount"], adj["quantity"],
                        ])
                    for svc in c["service_lines"]:
                        for adj in svc["adjustments"]:
                            adj_rows.append([
                                filename,
                                c["patient_control_number"],
                                f"Service ({svc['procedure_code']})",
                                adj["group_code"], adj["group_description"],
                                adj["reason_code"], adj["reason_description"],
                                adj["amount"], adj["quantity"],
                            ])
        write_sheet(ws_adj, adj_headers, adj_rows, currency_cols=[8])

    # -------------------------------------------------------------------
    # 837 sheets
    # -------------------------------------------------------------------
    if all_837:
        # --- 837 Claims ---
        if first_sheet:
            ws_claims = wb.active
            ws_claims.title = "837 Claims"
            first_sheet = False
        else:
            ws_claims = wb.create_sheet("837 Claims")

        claim_headers = [
            "Source File",
            "Claim ID", "Total Charges", "Place of Service",
            "Patient Name", "Patient DOB", "Patient Gender",
            "Subscriber Name", "Subscriber ID",
            "Payer Name", "Payer ID",
            "Billing Provider", "Billing Provider NPI", "Billing Provider Tax ID",
            "Rendering Provider", "Rendering Provider NPI",
            "Referring Provider",
            "Service Date From", "Service Date To",
            "Diagnosis Codes",
            "Prior Authorization",
            "Number of Service Lines",
        ]
        claim_rows = []
        for filename, transactions in all_837:
            for txn in transactions:
                for c in txn["claims"]:
                    dx_codes = ", ".join(
                        f"{d['code']} ({d['type']})" for d in c["diagnosis_codes"]
                    )
                    claim_rows.append([
                        filename,
                        c["claim_id"], c["total_charge"], c["place_of_service"],
                        c["patient_name"], c["patient_dob"], c["patient_gender"],
                        c["subscriber_name"], c["subscriber_id"],
                        c["payer_name"], c["payer_id"],
                        c["billing_provider_name"], c["billing_provider_npi"],
                        c["billing_provider_tax_id"],
                        c["rendering_provider_name"], c["rendering_provider_npi"],
                        c["referring_provider_name"],
                        c["service_date_from"], c["service_date_to"],
                        dx_codes,
                        c["prior_authorization"],
                        len(c["service_lines"]),
                    ])
        write_sheet(ws_claims, claim_headers, claim_rows, currency_cols=[3])

        # --- 837 Service Lines ---
        ws_svc = wb.create_sheet("837 Service Lines")
        svc_headers = [
            "Source File",
            "Claim ID", "Line #", "Procedure Code", "Modifiers",
            "Charge Amount", "Units", "Unit Type",
            "Place of Service", "Revenue Code",
            "Service Date From", "Service Date To",
            "Diagnosis Pointers", "NDC Code",
        ]
        svc_rows = []
        for filename, transactions in all_837:
            for txn in transactions:
                for c in txn["claims"]:
                    for svc in c["service_lines"]:
                        svc_rows.append([
                            filename,
                            svc.get("_claim_id", c["claim_id"]),
                            svc.get("line_number", ""),
                            svc["procedure_code"], svc["modifiers"],
                            svc["charge_amount"], svc["units"], svc["unit_type"],
                            svc.get("place_of_service", ""), svc.get("revenue_code", ""),
                            svc.get("service_date_from", ""), svc.get("service_date_to", ""),
                            svc.get("diagnosis_pointers", ""), svc.get("ndc_code", ""),
                        ])
        write_sheet(ws_svc, svc_headers, svc_rows, currency_cols=[6])

        # --- 837 Diagnosis Codes ---
        ws_dx = wb.create_sheet("837 Diagnosis Codes")
        dx_headers = [
            "Source File",
            "Claim ID", "Diagnosis Code", "Type", "Qualifier",
        ]
        dx_rows = []
        for filename, transactions in all_837:
            for txn in transactions:
                for c in txn["claims"]:
                    for dx in c["diagnosis_codes"]:
                        dx_rows.append([
                            filename,
                            c["claim_id"], dx["code"], dx["type"], dx["qualifier"],
                        ])
        write_sheet(ws_dx, dx_headers, dx_rows)

    wb.save(output_path)
    return output_path
